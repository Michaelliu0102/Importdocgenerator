"""
Excel模板填充模块
用于填充合同模板(CONTRACT_CAMARI_PRETTY.xlsx / CONTRACT.xlsx)和报关单模板(FedEx报关单模板.xlsx)
"""

import re
from typing import Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from config_loader import ConfigLoader
from item_declaration_mapper import (
    DEFAULT_IMPORT_EXCEL,
    load_excel_mapping_rows,
    load_import_rules_for_config,
    resolve_item_declaration,
    use_item_mapping_enabled,
)
from packing_slip_parser import country_name_to_cn
from pdf_parser import (
    strip_camari_bill_to_name_suffix,
    strip_ship_to_leaked_japan_from_camari_address,
)

CAMARI_BUYER_NAME = "CAMARI TRADING (ZHEJIANG) CO., LTD"
CAMARI_BUYER_ADDRESS = "1525 Hexing Rd, Jiaxing, Zhejiang, 314001, China"


# 含半角 ¥ (U+00A5) 与全角 ￥ (U+FFE5)，PDF 常混用
_CURRENCY_OR_AMOUNT = re.compile(
    r"[\u00a5\uffe5€]\s*[\d,]+(?:\.\d+)?|"
    r"\$\s*[\d,]+(?:\.\d+)?|"
    r"\b[\d,]+(?:\.\d+)?\s*(?:JPY|EUR|USD|CNY|GBP)\b",
    re.IGNORECASE,
)


def _clean_address_line(line: str) -> str:
    """去掉行尾/行内误提取的金额与币种，合并重复单词（如 Japan Japan）。"""
    s = (line or "").strip()
    if not s:
        return ""
    s = _CURRENCY_OR_AMOUNT.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    words = s.split()
    out: list[str] = []
    for w in words:
        if out and out[-1].lower() == w.lower():
            continue
        out.append(w)
    return " ".join(out)


def _dedupe_side_by_side_text(s: str) -> str:
    """PDF 双栏提取时常见「同一段文字并排重复」，去掉后半截重复。"""
    s = (s or "").strip()
    if not s:
        return ""
    parts = re.split(r"\s{2,}", s)
    if len(parts) == 2 and parts[0].strip() == parts[1].strip():
        return parts[0].strip()
    words = s.split()
    if len(words) >= 4 and len(words) % 2 == 0:
        h = len(words) // 2
        if words[:h] == words[h:]:
            return " ".join(words[:h])
    m = re.match(r"^(.+?)\s+(\1)$", s)
    if m:
        return m.group(1).strip()
    mid = len(s) // 2
    if mid > 5 and s[:mid].strip() == s[mid:].strip():
        return s[:mid].strip()
    return s


def _append_fedex_import_print_summary(wb) -> None:
    """在进口 FedEx 报关单工作簿中插入「打印汇总」工作表（置于首位），纵向拼接
    1.企业信息、2.商品信息、以及模板中若存在的「确认申报」表，便于只选该表一次打印。"""
    required = ("1.企业信息", "2.商品信息")
    if not all(name in wb.sheetnames for name in required):
        return

    SUMMARY = "打印汇总"
    third_candidates = ("3.确认申报", "确认申报", "3.确认")

    if SUMMARY in wb.sheetnames:
        wb.remove(wb[SUMMARY])
    ws_out = wb.create_sheet(SUMMARY, 0)

    row = 1
    hint = ws_out.cell(row=row, column=1)
    hint.value = "以下为「企业信息 + 商品信息 + 确认申报」汇总，便于一次打印；详细版式仍以各分表为准。"
    hint.font = Font(italic=True, size=9)
    row += 2

    def append_block(title: str, src) -> None:
        nonlocal row
        t = ws_out.cell(row=row, column=1)
        t.value = title
        t.font = Font(bold=True, size=11)
        row += 1
        mr = max(int(src.max_row or 0), 1)
        mc = max(int(src.max_column or 0), 1)
        for src_r in range(1, mr + 1):
            for c in range(1, mc + 1):
                ws_out.cell(row=row, column=c).value = src.cell(
                    row=src_r, column=c
                ).value
            row += 1
        row += 1

    append_block("一、企业信息", wb["1.企业信息"])
    append_block("二、商品信息", wb["2.商品信息"])
    for name in third_candidates:
        if name in wb.sheetnames:
            append_block("三、确认申报", wb[name])
            break


def _dedupe_party_field(s: str) -> str:
    if not (s or "").strip():
        return ""
    lines = []
    for line in (s or "").split("\n"):
        ln = line.strip()
        if not ln:
            continue
        ln = _clean_address_line(ln)
        if not ln:
            continue
        lines.append(_dedupe_side_by_side_text(ln))
    return "\n".join(lines)


def _effective_invoice_no(party: Dict[str, Any]) -> str:
    """合同号 / 发票号：仅使用发票解析的 invoice_no（不用装箱单号）。"""
    return (party.get("invoice_no") or "").strip()


def _buyer_country_last_line(invoice_data: Dict[str, Any]) -> str:
    """Bill To 地址中用于贸易国的行：自末行向上取首条非空清洗后的文本（通常为英文国家名）。"""
    ba = (invoice_data.get("buyer_address") or "").strip()
    if not ba:
        return ""
    for line in reversed([ln.strip() for ln in ba.split("\n") if ln.strip()]):
        cleaned = _clean_address_line(line)
        if cleaned:
            return cleaned
    return ""


def _find_cell_contains(ws, text: str):
    """返回第一个包含 text 的单元格（不区分大小写）。"""
    t = text.lower()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and t in str(cell.value).lower():
                return cell
    return None


def _unit_to_english(unit: str) -> str:
    mapping = {
        "米": "M",
        "张": "PCS",
        "平方米": "SQM",
        "平方英尺": "SQFT",
        "千克": "KG",
    }
    return mapping.get(unit, unit or "")


def _to_float(value):
    """将字符串/数字安全转换为 float，失败返回 None。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def _calc_gram_weight(net_weight, quantity):
    """
    计算每平方米克重(g/m2): net_weight / (quantity * 1.42) * 1000
    说明：按用户要求公式计算，net_weight 与 quantity 使用同一票据口径。
    """
    nw = _to_float(net_weight)
    qty = _to_float(quantity)
    if nw is None or qty is None or qty == 0:
        return None
    result = nw / (qty * 1.42) * 1000
    if result <= 0:
        return None
    return round(result, 2)


class ExcelFiller:
    def __init__(self, template_path: str):
        self.template_path = template_path

    def fill_contract_template(
        self,
        invoice_data: Dict[str, Any],
        supplier_info: Dict[str, Any],
        product_info: Dict[str, Any],
        output_path: str
    ):
        wb = load_workbook(self.template_path)
        ws = wb.active

        title = str(ws.cell(row=2, column=2).value or "").strip().upper()
        if title in {"CAMARI IMPORT CONTRACT", "CONTRACT"}:
            self._fill_contract_pretty(ws, invoice_data, supplier_info, product_info)
        else:
            self._fill_contract_legacy(ws, invoice_data, supplier_info, product_info)

        wb.save(output_path)
        return output_path

    def _fill_contract_pretty(
        self,
        ws,
        invoice_data: Dict[str, Any],
        supplier_info: Dict[str, Any],
        product_info: Dict[str, Any],
    ):
        """填充新版排版合同模板 CONTRACT_CAMARI_PRETTY.xlsx。"""
        items = invoice_data.get("items", [])
        supplier_name = supplier_info.get("name", invoice_data.get("supplier_name", ""))
        supplier_addr = supplier_info.get("address", "")
        trade_term = invoice_data.get("trade_term") or supplier_info.get("trade_term", "")
        currency = invoice_data.get("currency", "EUR") or "EUR"
        invoice_no = invoice_data.get("invoice_no", "")
        invoice_date = invoice_data.get("invoice_date", "")
        payment = invoice_data.get("payment_cond") or supplier_info.get("payment_term", "")
        is_camari_cust = invoice_data.get("format") == "camari_cust"

        # Header block：CustInvc 内销发票 = 卖方 issuer（日本/境外）+ Bill To 买方（嘉兴）
        if is_camari_cust:
            # Row 4 = Buyer (固定), Row 7 = Seller (PDF issuer)
            ws.cell(row=4, column=3).value = CAMARI_BUYER_NAME
            ws.cell(row=5, column=3).value = CAMARI_BUYER_ADDRESS
            ws.cell(row=7, column=3).value = (
                invoice_data.get("issuer_name")
                or "CAMARI INTERNATIONAL JAPAN"
            )
            ws.cell(row=8, column=3).value = invoice_data.get("issuer_address") or ""
        else:
            ws.cell(row=4, column=3).value = CAMARI_BUYER_NAME
            ws.cell(row=7, column=3).value = supplier_name
            ws.cell(row=8, column=3).value = supplier_addr
        ws.cell(row=4, column=10).value = invoice_no
        ws.cell(row=5, column=10).value = invoice_date
        # 第10行已改为只保留 Incoterms / Payment Terms（无 Country / Currency）
        ws.cell(row=10, column=3).value = trade_term
        ws.cell(row=10, column=7).value = payment

        # Item table rows: 13-24（No. = 发票 # 栏行号；品名 = Item 描述）
        for i, item in enumerate(items[:12], 1):
            r = 12 + i
            row_no = item.get("line_no")
            if row_no is not None and str(row_no).strip() != "":
                try:
                    ws.cell(row=r, column=2).value = int(str(row_no).strip())
                except (TypeError, ValueError):
                    ws.cell(row=r, column=2).value = row_no
            else:
                ws.cell(row=r, column=2).value = i
            ws.cell(row=r, column=3).value = item.get("description", "")

            qty_val = item.get("quantity", "")
            try:
                qty_val = float(qty_val)
            except (TypeError, ValueError):
                pass
            ws.cell(row=r, column=7).value = qty_val
            ws.cell(row=r, column=8).value = _unit_to_english(item.get("unit", "M"))

            up_val = item.get("unit_price", "")
            try:
                up_val = float(up_val)
            except (TypeError, ValueError):
                pass
            ws.cell(row=r, column=9).value = up_val

            ws.cell(row=r, column=10).value = currency

            amt_val = item.get("amount", "")
            try:
                amt_val = float(amt_val)
            except (TypeError, ValueError):
                pass
            ws.cell(row=r, column=11).value = amt_val

        # Total
        total = invoice_data.get("total_amount", "")
        try:
            total = float(total)
        except (TypeError, ValueError):
            pass
        ws.cell(row=25, column=11).value = total

    def _fill_contract_legacy(
        self,
        ws,
        invoice_data: Dict[str, Any],
        supplier_info: Dict[str, Any],
        product_info: Dict[str, Any],
    ):
        """兼容旧版 CONTRACT.xlsx 的填充逻辑。"""

        items = invoice_data.get("items", [])
        supplier_name = supplier_info.get("name", invoice_data.get("supplier_name", ""))
        trade_term = invoice_data.get("trade_term") or supplier_info.get("trade_term", "")
        currency = invoice_data.get("currency", "EUR") or "EUR"
        invoice_no = invoice_data.get("invoice_no", "")
        invoice_date = invoice_data.get("invoice_date", "")
        payment = invoice_data.get("payment_cond") or supplier_info.get("payment_term", "")

        # 买方固定（与模板一致）
        ws.cell(row=4, column=3).value = "CAMARI TRADING (ZHEJIANG) CO., LTD"

        # 合同编号 = Invoice number
        ws.cell(row=5, column=11).value = invoice_no

        ws.cell(row=7, column=3).value = supplier_name
        ws.cell(row=7, column=11).value = invoice_date

        # 第12行：货币 / Incoterms / Payment（CAMARI 模板）
        b12 = ws.cell(row=12, column=2).value
        if b12 and "currency" in str(b12).lower():
            ws.cell(row=12, column=3).value = currency
            ws.cell(row=12, column=6).value = trade_term or ""
            ws.cell(row=12, column=9).value = payment or ""
        else:
            # 旧模板：贸易术语写在首行单价列旁
            if trade_term and items:
                ws.cell(row=15, column=10).value = trade_term

        # 表头币种
        hdr = _find_cell_contains(ws, "name of commodity")
        header_row = hdr.row if hdr else 14
        ws.cell(row=header_row, column=10).value = f"Unit Price {currency}"
        ws.cell(row=header_row, column=11).value = f"Amount {currency}"

        first_item_row = header_row + 1

        for i, item in enumerate(items):
            r = first_item_row + i
            ws.cell(row=r, column=3).value = item.get("description", "")

            try:
                ws.cell(row=r, column=8).value = float(item.get("quantity", 0))
            except (ValueError, TypeError):
                ws.cell(row=r, column=8).value = item.get("quantity", "")

            try:
                ws.cell(row=r, column=10).value = float(item.get("unit_price", 0))
            except (ValueError, TypeError):
                ws.cell(row=r, column=10).value = item.get("unit_price", "")

            try:
                ws.cell(row=r, column=11).value = float(item.get("amount", 0))
            except (ValueError, TypeError):
                ws.cell(row=r, column=11).value = item.get("amount", "")

        try:
            total = float(invoice_data.get("total_amount", 0))
        except (ValueError, TypeError):
            total = invoice_data.get("total_amount", "")

        tot_cell = _find_cell_contains(ws, "TOT AMOUNT")
        if tot_cell is not None:
            tot_cell.value = f"TOT AMOUNT: {total} {currency}".strip()

        place = _find_cell_contains(ws, "Place of Shipment")
        if place is not None:
            country = supplier_info.get("country", "")
            if country:
                ws.cell(row=place.row, column=4).value = f"From {country}"

        pay_clause = _find_cell_contains(ws, "(3) Terms of payment")
        if pay_clause is not None and payment:
            ws.cell(row=pay_clause.row, column=5).value = payment

    def fill_customs_declaration_template(
        self,
        invoice_data: Dict[str, Any],
        supplier_info: Dict[str, Any],
        product_info: Dict[str, Any],
        output_path: str,
        product_info_map: Dict[str, Dict] = None,
        config_path: str = None,
    ):
        """
        填充FedEx报关单模板 (FedEx报关单模板.xlsx)

        product_info_map: optional {hide_type: product_info} for per-item HS/elements.
        """
        wb = load_workbook(self.template_path)

        items = invoice_data.get("customs_items") or invoice_data.get("items", [])
        currency = invoice_data.get("currency", "EUR")
        trade_term = invoice_data.get("trade_term", supplier_info.get("trade_term", ""))
        supplier_country = supplier_info.get("country", "")
        net_weight = invoice_data.get("net_weight", "")

        if not product_info_map:
            product_info_map = {}

        mapping_loader = None
        import_rules = []
        internal_item_rows = []
        if config_path and use_item_mapping_enabled(invoice_data):
            mapping_loader = ConfigLoader(config_path)
            import_rules = load_import_rules_for_config(config_path, invoice_data)
            internal_item_rows = load_excel_mapping_rows(DEFAULT_IMPORT_EXCEL)

        # === Sheet 1: 企业信息 ===
        if "1.企业信息" in wb.sheetnames:
            ws1 = wb["1.企业信息"]
            ws1.cell(row=9, column=2).value = invoice_data.get("invoice_no", "")

        # === Sheet 2: 商品信息 ===
        if "2.商品信息" in wb.sheetnames:
            ws2 = wb["2.商品信息"]

            item_count = len(items) if items else 1
            per_item_weight = ""
            if net_weight:
                try:
                    per_item_weight = f"{float(net_weight) / item_count:.2f}"
                except (ValueError, TypeError):
                    per_item_weight = net_weight

            for i, item in enumerate(items):
                r = 2 + i
                qty = item.get("quantity", "")

                if mapping_loader is not None:
                    resolved = resolve_item_declaration(
                        item,
                        [],
                        mapping_loader,
                        product_info or {},
                        product_info_map or {},
                        import_rules=import_rules,
                        invoice_data=invoice_data,
                        internal_item_rows=internal_item_rows,
                    )
                    hs_code = resolved.get("hs_code", "")
                    product_name_cn = resolved.get("product_name", "")
                    decl_elements = resolved.get("declaration_elements", {})
                else:
                    # Pick product_info for this item: by hide_type, then default
                    hide_type = item.get("hide_type", "")
                    cur_pi = product_info_map.get(hide_type) or product_info or {}
                    hs_code = cur_pi.get("hs_code", "")
                    product_name_cn = cur_pi.get("name", "")
                    decl_elements = cur_pi.get("declaration_elements", {})

                gram_weight = _calc_gram_weight(net_weight, qty)

                item_decl_elements = dict(decl_elements)
                for k, v in list(item_decl_elements.items()):
                    key_has_gram = ("克重" in str(k)) or ("g/m2" in str(k).lower())
                    if not key_has_gram:
                        continue
                    val = "" if v is None else str(v).strip()
                    if (not val or val == "克/平方米") and gram_weight is not None:
                        item_decl_elements[k] = f"{gram_weight} g/m2"

                decl_parts = []
                for idx, (k, v) in enumerate(item_decl_elements.items(), 1):
                    if v:
                        decl_parts.append(f"{idx}:{k}:{v}")
                decl_text = "; ".join(decl_parts) if decl_parts else ""

                prefix = item.get("item_code_prefix", "")
                if prefix == "5012":
                    cn_name = "汽车装饰用材料"
                    item_hs = "5603149000"
                elif prefix in ("5015", "5030", "5010"):
                    cn_name = "面料"
                    item_hs = "5603149000"
                else:
                    cn_name = product_name_cn
                    item_hs = hs_code

                ws2.cell(row=r, column=1).value = cn_name
                ws2.cell(row=r, column=2).value = item_hs
                c3 = ws2.cell(row=r, column=3)
                c3.value = decl_text
                # C 列申报要素：自动换行（依列宽折行）；顶对齐便于多行阅读
                c3.alignment = Alignment(wrap_text=True, vertical="top")
                if decl_text:
                    # openpyxl 不会按内容重算行高，粗算行数避免打开前被裁切（Excel 仍可再调）
                    est_lines = max(1, len(decl_text) // 42 + decl_text.count(";") // 2 + 1)
                    ws2.row_dimensions[r].height = min(409, 14 * est_lines)

                ws2.cell(row=r, column=4).value = per_item_weight

                unit = item.get("unit", "米")
                ws2.cell(row=r, column=5).value = f"{qty} {unit}"

                origin = _country_code_to_chinese(item.get("country_code", ""))
                if not origin:
                    origin = supplier_country
                ws2.cell(row=r, column=6).value = origin

                amount = item.get("amount", "")
                ws2.cell(row=r, column=7).value = f"{currency} {amount}"

                ws2.cell(row=r, column=8).value = trade_term
                ws2.cell(row=r, column=9).value = supplier_country

        _append_fedex_import_print_summary(wb)

        wb.save(output_path)
        return output_path

    def fill_export_contract_template(
        self,
        invoice_data: Dict[str, Any],
        supplier_info: Dict[str, Any],
        product_info: Dict[str, Any],
        output_path: str,
        *,
        party_invoice_data: Optional[Dict[str, Any]] = None,
    ) -> str:
        """填充 export_templates/export_contract.xlsx（出口合同，版式与进口 CONTRACT 不同）。

        party_invoice_data: 若提供，买方/卖方抬头与合同号取自**原票**解析结果（与 EUR 换算后的
        invoice_data 分离）；行项目金额仍以 invoice_data 为准。合同号仅使用 invoice_no。
        """
        wb = load_workbook(self.template_path)
        ws = wb.active

        p = party_invoice_data if party_invoice_data is not None else invoice_data
        items = invoice_data.get("items", [])
        buyer = CAMARI_BUYER_NAME
        buyer_addr = CAMARI_BUYER_ADDRESS
        issuer_name = _dedupe_party_field((p.get("issuer_name") or "").strip())
        issuer_addr = _dedupe_party_field((p.get("issuer_address") or "").strip())
        invoice_no = _effective_invoice_no(p)
        invoice_date = (p.get("invoice_date") or "").strip()
        trade_term = invoice_data.get("trade_term") or (
            supplier_info or {}
        ).get("trade_term", "")
        payment = invoice_data.get("payment_cond") or (
            supplier_info or {}
        ).get("payment_term", "")
        currency = invoice_data.get("currency", "EUR") or "EUR"

        ws["C4"] = buyer
        ws["C5"] = buyer_addr
        ws["J4"] = invoice_no
        ws["J5"] = invoice_date
        ws["C10"] = trade_term
        ws["G10"] = payment
        # 卖方（原发票 issuer；若模板 C2/C3 为占位或空则写入）
        if issuer_name or issuer_addr:
            for r, val in ((2, issuer_name), (3, issuer_addr)):
                if not val:
                    continue
                cur = ws.cell(row=r, column=3).value
                if cur is None or str(cur).strip() == "":
                    ws.cell(row=r, column=3).value = val
            lab = _find_cell_contains(ws, "卖方")
            if lab and issuer_name and "买方" not in str(lab.value or ""):
                ws.cell(row=lab.row, column=lab.column + 1).value = issuer_name
            lab_addr = _find_cell_contains(ws, "卖方地址")
            if lab_addr and issuer_addr:
                ws.cell(row=lab_addr.row, column=lab_addr.column + 1).value = issuer_addr

        for i, item in enumerate(items[:12], 1):
            r = 12 + i
            row_no = item.get("line_no")
            if row_no is not None and str(row_no).strip() != "":
                try:
                    ws.cell(row=r, column=2).value = int(str(row_no).strip())
                except (TypeError, ValueError):
                    ws.cell(row=r, column=2).value = row_no
            else:
                ws.cell(row=r, column=2).value = i
            ws.cell(row=r, column=3).value = item.get("description", "")

            qty_val = item.get("quantity", "")
            try:
                qty_val = float(qty_val)
            except (TypeError, ValueError):
                pass
            ws.cell(row=r, column=7).value = qty_val
            ws.cell(row=r, column=8).value = _unit_to_english(item.get("unit", "M"))

            up_val = item.get("unit_price", "")
            try:
                up_val = float(up_val)
            except (TypeError, ValueError):
                pass
            ws.cell(row=r, column=9).value = up_val
            ws.cell(row=r, column=10).value = currency

            amt_val = item.get("amount", "")
            try:
                amt_val = float(amt_val)
            except (TypeError, ValueError):
                pass
            ws.cell(row=r, column=11).value = amt_val

        total = invoice_data.get("total_amount", "")
        try:
            total = float(total)
        except (TypeError, ValueError):
            pass
        ws["K25"] = total

        wb.save(output_path)
        return output_path

    def fill_export_declaration_template(
        self,
        invoice_data: Dict[str, Any],
        supplier_info: Dict[str, Any],
        product_info: Dict[str, Any],
        output_path: str,
        line_groups: list,
        product_info_map: Dict[str, Dict] = None,
        packing_slip_data: Optional[Dict[str, Any]] = None,
        *,
        party_invoice_data: Optional[Dict[str, Any]] = None,
    ) -> str:
        """填充 export_templates/export_declaration.xlsx（海关出口报关单样式）。

        party_invoice_data: 若提供，境外收货人抬头与合同号取自**原票**解析结果；金额行仍以
        invoice_data / line_groups（可为 EUR 换算后）为准。
        """
        wb = load_workbook(self.template_path)
        ws = wb.active

        if product_info_map is None:
            product_info_map = {}

        p = party_invoice_data if party_invoice_data is not None else invoice_data
        buyer = CAMARI_BUYER_NAME
        buyer_addr = CAMARI_BUYER_ADDRESS
        block = "境外收货人\n"
        if buyer:
            block += buyer
        if buyer_addr:
            block += "\n" + buyer_addr
        ws["A3"] = block

        inv_no = _effective_invoice_no(p)
        trade_term = (
            invoice_data.get("trade_term")
            or (supplier_info or {}).get("trade_term", "")
            or ""
        ).strip()
        bill_country_cn = country_name_to_cn(_buyer_country_last_line(p))
        dest_cn = ""
        if packing_slip_data:
            dest_cn = (
                packing_slip_data.get("ship_to_country_cn")
                or country_name_to_cn(packing_slip_data.get("ship_to_country") or "")
            )

        if inv_no:
            ws["A5"] = f"合同号\n{inv_no}"
        if bill_country_cn:
            ws["C5"] = f"贸易国（地区）\n{bill_country_cn}"
        if dest_cn:
            ws["E5"] = f"运抵国（地区）\n{dest_cn}"

        if packing_slip_data:
            pkg = packing_slip_data.get("pkg_qty")
            gw = packing_slip_data.get("gross_weight_kg")
            nw = packing_slip_data.get("net_weight_kg")
            if pkg is not None and str(pkg).strip() != "":
                ws["C6"] = f"件数\n{pkg}"
            if gw is not None and str(gw).strip() != "":
                ws["D6"] = f"毛重（千克）\n{gw}"
            if nw is not None and str(nw).strip() != "":
                ws["E6"] = f"净重(千克)\n{nw}"
        if trade_term:
            ws["F6"] = f"成交方式\n{trade_term}"

        currency = invoice_data.get("currency", "") or "EUR"

        for idx, grp in enumerate(line_groups[:6]):
            r = 10 + idx
            if r >= 16:
                break
            hs = grp.get("hs_code", "")
            pname = grp.get("product_name", "")
            gitems = grp.get("items") or []
            desc_parts = []
            total_qty = 0.0
            total_amt = 0.0
            unit = ""
            for it in gitems:
                d = (it.get("description") or "").strip()
                if d:
                    desc_parts.append(d)
                unit = it.get("unit") or unit
                try:
                    total_qty += float(it.get("quantity", 0))
                except (TypeError, ValueError):
                    pass
                try:
                    total_amt += float(it.get("amount", 0))
                except (TypeError, ValueError):
                    pass
            desc = pname or "; ".join(desc_parts[:3])
            if len(desc_parts) > 3:
                desc += "…"
            u_en = _unit_to_english(unit or (gitems[0].get("unit") if gitems else "M") or "M")
            qty_unit = f"{total_qty:g} {u_en}" if gitems else ""

            price_line = ""
            if gitems:
                if total_qty > 0 and total_amt > 0:
                    try:
                        up = round(total_amt / total_qty, 4)
                    except (TypeError, ZeroDivisionError):
                        up = None
                else:
                    up = None
                if up is None:
                    try:
                        up = float(gitems[0].get("unit_price", 0))
                    except (TypeError, ValueError):
                        up = gitems[0].get("unit_price", "")
                price_line = f"{up} {currency} / {total_amt} {currency}"

            ws.cell(row=r, column=1).value = idx + 1
            ws.cell(row=r, column=2).value = hs
            ws.cell(row=r, column=3).value = desc
            ws.cell(row=r, column=4).value = qty_unit
            ws.cell(row=r, column=5).value = price_line

        wb.save(output_path)
        return output_path


def _country_code_to_chinese(code: str) -> str:
    mapping = {
        "IT": "意大利",
        "DE": "德国",
        "GB": "英国",
        "US": "美国",
        "NL": "荷兰",
        "FR": "法国",
        "CN": "中国",
    }
    return mapping.get(code, "")
