"""
Invoice 行 description / item_code 与申报品名、申报要素的映射。
出口优先读取 export_templates/出口申报要素对应表.xlsx 的 ITEM、品名、HS Code、申报要素四列。
"""

from __future__ import annotations

import copy
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from config_loader import ConfigLoader

ROOT = Path(__file__).resolve().parent
DEFAULT_IMPORT_EXCEL = ROOT / "data" / "item和品名对应表.xlsx"
DEFAULT_IMPORT_RULES_EXCEL = ROOT / "data" / "进口商品识别规则.xlsx"
DEFAULT_EXPORT_EXCEL = ROOT / "export_templates" / "出口申报要素对应表.xlsx"
MappingRow = Dict[str, Any]

# 旧两列表 Excel 英文 ITEM 列与 YAML product_categories 键的对应
EXCEL_ITEM_TO_YAML_KEY: Dict[str, str] = {
    "Phone Case": "export_hs_392690_misc",
    "Airpods Case": "export_hs_392690_misc",
    "Watch Band": "export_hs_911390_watch",
    "Sunvisor Cover": "export_hs_392690_misc",
    "Pouch": "export_hs_392690_misc",
    "Lumbar Cushion": "export_hs_940199_auto",
    "Headrest Pillow": "export_hs_940199_auto",
    "Tray": "export_hs_392690_misc",
    "Bagpack": "export_hs_420212_luggage",
    "Luggage": "export_hs_420212_luggage",
    "Duffle Bag": "export_hs_420212_luggage",
    "Key Fob": "export_hs_392690_misc",
    "Card Holder": "export_hs_392690_misc",
    "Seat Cover": "alcantara_5015_5030_5010",
    "Notebook": "export_hs_482010_notebook",
    "Color Cards": "export_hs_491110_print",
    "Spectacle Case": "export_hs_392690_misc",
}


def _resolve_config_path(config_path: str) -> Path:
    p = Path(config_path)
    if not p.is_absolute():
        p = (Path.cwd() / p).resolve()
    else:
        p = p.resolve()
    return p


def _parse_declaration_elements(raw: Any, product_name: str) -> Dict[str, str]:
    """把 Excel D 列申报要素转成稳定的 dict；不强行理解无标签的管道格式。"""
    elements: Dict[str, str] = {}

    text = "" if raw is None else str(raw).strip()
    if product_name == "牛皮" and "|" in text:
        return _parse_cowhide_pipe_declaration(text, product_name)

    if product_name:
        elements["品名"] = product_name

    if not text:
        return elements

    normalized = (
        text.replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("；", "\n")
        .replace(";", "\n")
    )
    parts = [p.strip(" \t\n:：") for p in normalized.split("\n") if p.strip()]
    if len(parts) <= 1 and "|" in text:
        parts = [p.strip(" \t\n|") for p in text.split("|") if p.strip()]

    for idx, part in enumerate(parts, 1):
        key = ""
        value = ""
        if "：" in part:
            key, value = part.split("：", 1)
        elif ":" in part:
            key, value = part.split(":", 1)
        if key:
            key = key.strip()
            value = value.strip()
            elements[key or f"要素{idx}"] = value
        else:
            elements[f"要素{idx}"] = part.strip()
    return elements


def _parse_cowhide_pipe_declaration(raw: str, product_name: str) -> Dict[str, str]:
    """出口表牛皮行使用海关要素位序，转为可读字段名。"""
    parts = [p.strip(" \t\n\r；;") for p in raw.split("|")]

    def part(idx: int) -> str:
        return parts[idx].strip() if idx < len(parts) else ""

    color = ""
    color_usage = part(9)
    color_m = re.search(r"颜色[:：]?\s*([^。.\s]+)", color_usage)
    if color_m:
        color = color_m.group(1).strip()

    spec_bits = [part(5), part(7)]
    spec = "，".join(bit for bit in spec_bits if bit)

    return {
        "涂覆": part(2),
        "粒面层": part(3),
        "动物种类": part(4),
        "规格": spec,
        "颜色": color or "等",
        "用途": "用于家具包覆",
        "鞣制": part(12),
        "计量单位": part(13).strip(" ；;") or "张",
        "品名": product_name,
    }


def load_excel_mapping_rows(xlsx_path: Path) -> List[MappingRow]:
    """返回 Excel 映射行，兼容旧两列表和出口四列表。"""
    if not xlsx_path.exists():
        return []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: List[MappingRow] = []
        for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_no == 1:
                continue
            if not row or len(row) < 2:
                continue
            a, b = row[0], row[1]
            if a is None or b is None:
                continue
            item_en = str(a).strip()
            name_cn = str(b).strip()
            if not item_en or not name_cn:
                continue
            hs_code = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            raw_elements = row[3] if len(row) > 3 else None
            rows.append(
                {
                    "row_no": row_no,
                    "item": item_en,
                    "product_name": name_cn,
                    "hs_code": hs_code,
                    "declaration_elements": _parse_declaration_elements(
                        raw_elements, name_cn
                    ),
                }
            )
        return rows
    finally:
        wb.close()


def load_import_rule_rows(xlsx_path: Path) -> List[MappingRow]:
    """读取进口识别规则：供应商原始描述/料号 -> YAML 内部标准品类。"""
    if not xlsx_path.exists():
        return []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return []
        headers = {
            str(value).strip().lower(): idx
            for idx, value in enumerate(header)
            if value is not None and str(value).strip()
        }

        def cell(row: tuple, *names: str) -> str:
            for name in names:
                idx = headers.get(name.lower())
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
            return ""

        rules: List[MappingRow] = []
        for row_no, row in enumerate(rows_iter, start=2):
            enabled = cell(row, "启用", "是否启用", "enabled")
            if enabled and enabled.strip().lower() in {
                "否",
                "no",
                "n",
                "false",
                "0",
                "停用",
            }:
                continue

            pattern = cell(row, "匹配内容", "内容", "pattern", "match_value")
            target_item = cell(
                row,
                "内部ITEM",
                "内部ITEM / YAML品类",
                "内部item",
                "内部item / yaml品类",
                "内部品类",
                "标准品类",
                "product_key",
                "产品编码",
            )
            if not pattern or not target_item:
                continue

            priority_raw = cell(row, "优先级", "priority")
            try:
                priority = int(float(priority_raw)) if priority_raw else 0
            except ValueError:
                priority = 0

            rules.append(
                {
                    "row_no": row_no,
                    "enabled": True,
                    "supplier": cell(row, "供应商", "supplier", "supplier_code"),
                    "field": cell(
                        row, "匹配字段", "字段", "field", "match_field"
                    )
                    or "任意文本",
                    "match_type": cell(row, "匹配方式", "方式", "match_type") or "包含",
                    "pattern": pattern,
                    "target_item": target_item,
                    "product_key": target_item,
                    "product_name": cell(row, "申报品名", "品名", "product_name"),
                    "priority": priority,
                    "note": cell(row, "备注", "note"),
                }
            )
        return sorted(
            rules,
            key=lambda r: (-int(r.get("priority") or 0), int(r.get("row_no") or 0)),
        )
    finally:
        wb.close()


def load_import_rules_for_config(
    config_path: str,
    invoice_data: Optional[Dict[str, Any]] = None,
    mapping_xlsx: Optional[str] = None,
) -> List[MappingRow]:
    return load_import_rule_rows(
        _resolve_import_rules_xlsx_path(config_path, mapping_xlsx, invoice_data)
    )


def _match_excel_row(
    item: Dict[str, Any], excel_rows: List[MappingRow]
) -> Optional[MappingRow]:
    """按 ITEM 名称优先匹配，再按内容语义匹配特殊前缀规则。"""
    description = _item_desc_for_declaration(item)
    d = (description or "").strip().lower()
    if not excel_rows:
        return None

    # 普通 ITEM 名称：按英文 ITEM 子串匹配，取最长命中。
    best: Optional[MappingRow] = None
    best_len = -1
    if d:
        for row in excel_rows:
            item_en = str(row.get("item") or "").strip()
            ie = item_en.lower()
            if not ie:
                continue
            if ie in d and len(ie) > best_len:
                best = row
                best_len = len(ie)
    if best:
        return best

    candidates = _item_prefix_candidates(item)
    for row in excel_rows:
        raw_item = str(row.get("item") or "").strip()
        prefixes = [_normalize_prefix_token(p) for p in raw_item.split("/") if p]
        special_rule = _special_excel_prefix_rule(raw_item, prefixes)
        if special_rule == "four_digit_prefix":
            if any(_candidate_starts_with(candidates, p, 4) for p in prefixes):
                return row
        elif special_rule == "full_token_prefix":
            if any(_candidate_starts_with(candidates, p, len(p)) for p in prefixes):
                return row
        elif special_rule == "fa_prefix":
            if any(_candidate_starts_with(candidates, p, 6) for p in prefixes):
                return row
        elif special_rule == "two_char_prefix":
            if any(_candidate_starts_with(candidates, p, 2) for p in prefixes):
                return row
    return None


def _special_excel_prefix_rule(raw_item: str, prefixes: List[str]) -> Optional[str]:
    if not prefixes:
        return None

    if all(re.fullmatch(r"\d{4}", p) for p in prefixes):
        return "four_digit_prefix"

    if all(p.startswith("FA") and len(p) >= 6 for p in prefixes):
        return "fa_prefix"

    if len(prefixes) == 1 and prefixes[0] == "MF":
        return "two_char_prefix"

    if any(p in {"NAPPA", "VERONA", "ROMA"} for p in prefixes):
        return "full_token_prefix"

    return None


def _split_rule_tokens(value: Any, *, allow_slash: bool = True) -> List[str]:
    text = str(value or "").strip()
    if not text:
        return []
    normalized = (
        text.replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("；", ";")
        .replace("，", ",")
        .replace("、", ",")
    )
    if allow_slash and "/" in normalized and not re.search(r"\s/\s", normalized):
        normalized = normalized.replace("/", ",")
    parts = re.split(r"[\n;,]+", normalized)
    return [p.strip() for p in parts if p and p.strip()]


def _supplier_matches_rule(
    rule_supplier: Any,
    invoice_data: Dict[str, Any],
    loader: ConfigLoader,
) -> bool:
    tokens = _split_rule_tokens(rule_supplier, allow_slash=False)
    if not tokens:
        return True

    supplier_code = str(invoice_data.get("supplier_code") or "").strip().lower()
    supplier_name = str(invoice_data.get("supplier_name") or "").strip().lower()

    for token in tokens:
        tl = token.lower()
        if tl in {"*", "all", "全部"}:
            return True
        if supplier_code and tl == supplier_code:
            return True
        if supplier_name and tl in supplier_name:
            return True

        info = loader.get_supplier_info(token)
        cfg_name = str((info or {}).get("name") or "").strip().lower()
        if cfg_name and supplier_name and (
            cfg_name in supplier_name or supplier_name in cfg_name
        ):
            return True
    return False


def _field_values_for_rule(item: Dict[str, Any], field: Any) -> List[str]:
    f = str(field or "").strip().lower()
    desc = _item_desc_for_declaration(item)
    first_four = _first_four_digits(item.get("item_code")) or ""
    groups = {
        "item_code": [item.get("item_code"), first_four],
        "料号": [item.get("item_code"), first_four],
        "item_code_prefix": [item.get("item_code_prefix"), first_four],
        "料号前缀": [item.get("item_code_prefix"), first_four],
        "description": [desc, item.get("description"), item.get("description_supplement")],
        "描述": [desc, item.get("description"), item.get("description_supplement")],
        "composition": [item.get("composition"), desc],
        "成分": [item.get("composition"), desc],
        "hide_type": [item.get("hide_type")],
        "皮料类型": [item.get("hide_type")],
        "article_name": [item.get("article_name")],
        "款号": [item.get("article_name")],
    }

    if f in {"", "any", "任意", "任意文本", "all", "全文"}:
        values = [
            item.get("item_code"),
            item.get("item_code_prefix"),
            item.get("article_name"),
            item.get("description"),
            item.get("description_supplement"),
            item.get("composition"),
            item.get("hide_type"),
            desc,
            first_four,
        ]
    else:
        values = groups.get(f, [item.get(field)])

    out: List[str] = []
    for value in values:
        s = str(value or "").strip()
        if s and s not in out:
            out.append(s)
    return out


def _rule_pattern_matches(values: List[str], match_type: Any, pattern: Any) -> bool:
    mt = str(match_type or "包含").strip().lower()
    patterns = (
        [str(pattern or "").strip()]
        if mt in {"regex", "regexp", "正则"}
        else _split_rule_tokens(pattern)
    )
    if not values or not patterns:
        return False

    for raw_pat in patterns:
        pat = raw_pat.strip()
        if not pat:
            continue
        pat_l = pat.lower()
        pat_norm = _normalize_prefix_token(pat)

        for value in values:
            val = value.strip()
            val_l = val.lower()
            val_norm = _normalize_prefix_token(val)

            if mt in {"等于", "=", "==", "exact", "equals", "equal"}:
                if val_l == pat_l or (pat_norm and val_norm == pat_norm):
                    return True
            elif mt in {"前缀", "prefix", "starts_with", "startswith"}:
                if val_l.startswith(pat_l) or (
                    pat_norm and val_norm.startswith(pat_norm)
                ):
                    return True
            elif mt in {"正则", "regex", "regexp"}:
                try:
                    if re.search(pat, val, re.IGNORECASE):
                        return True
                except re.error:
                    continue
            else:
                if pat_l in val_l or (pat_norm and pat_norm in val_norm):
                    return True
    return False


def _match_import_rule(
    item: Dict[str, Any],
    import_rules: List[MappingRow],
    invoice_data: Dict[str, Any],
    loader: ConfigLoader,
) -> Optional[MappingRow]:
    for rule in import_rules:
        if not _supplier_matches_rule(rule.get("supplier"), invoice_data, loader):
            continue
        values = _field_values_for_rule(item, rule.get("field"))
        if _rule_pattern_matches(values, rule.get("match_type"), rule.get("pattern")):
            return rule
    return None


def _normalize_prefix_token(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _strip_line_number(value: Any) -> str:
    # CustInvc 行项目常见 "1MF ..." / "4VERONA ..." / "1 NAPPA ..."，
    # 但 50153335BP 这种数字料号不能被剥掉。
    return re.sub(r"^\s*\d{1,3}(?=\s*[A-Za-z])\s*", "", str(value or "").strip())


def _item_prefix_candidates(item: Dict[str, Any]) -> List[str]:
    values = [
        item.get("item_code"),
        item.get("article_name"),
        item.get("description"),
        item.get("description_supplement"),
        _item_desc_for_declaration(item),
    ]
    candidates: List[str] = []
    for value in values:
        raw = _strip_line_number(value)
        norm = _normalize_prefix_token(raw)
        if norm and norm not in candidates:
            candidates.append(norm)
        # 聚合后的皮料行可能是 "LEATHER ART. NAPPA ..."，也要让 ART 后的货号参与判断。
        art = re.search(r"\bART\.?\s+([A-Z0-9]+)", raw, re.IGNORECASE)
        if art:
            norm_art = _normalize_prefix_token(art.group(1))
            if norm_art and norm_art not in candidates:
                candidates.append(norm_art)
    return candidates


def _candidate_starts_with(
    candidates: List[str], prefix: str, significant_len: int
) -> bool:
    p = _normalize_prefix_token(prefix)[:significant_len]
    return bool(p) and any(c.startswith(p) for c in candidates)


def _first_four_digits(item_code: Any) -> Optional[str]:
    if item_code is None:
        return None
    digits = re.sub(r"\D", "", str(item_code).strip())
    if len(digits) >= 4:
        return digits[:4]
    return None


def _leading_after_line_digits(s: str) -> str:
    """去掉行首数字编号后用于 MF / 皮革前缀判断。"""
    t = (s or "").strip()
    t = re.sub(r"^\d+[\s\-]*", "", t)
    return t.strip()


def _mf_prefix(description: str) -> bool:
    lead = _leading_after_line_digits(description)
    return len(lead) >= 2 and lead[:2].upper() == "MF"


def _leather_keyword_prefix(description: str) -> bool:
    """NAPPA / VERONA / ROMA 为去掉行首数字后的字母前缀（大小写不敏感）。

    兼容 "LEATHER ART. NAPPA ..." / "ART. VERONA ..." 这类前置标签。
    """
    s = _leading_after_line_digits(description)
    if not s:
        return False
    u = s.upper()
    u = re.sub(r"^(?:LEATHER\s+)?ART\.?\s+", "", u).strip()
    if len(u) >= 5 and u.startswith("NAPPA"):
        return True
    if len(u) >= 6 and u.startswith("VERONA"):
        return True
    if len(u) >= 4 and u.startswith("ROMA"):
        return True
    return False


def _product_block_from_yaml(
    loader: ConfigLoader, key: str
) -> Dict[str, Any]:
    pi = loader.get_product_info(key) or {}
    return {
        "hs_code": (pi.get("hs_code") or "").strip(),
        "product_name": (pi.get("name") or "").strip(),
        "declaration_elements": copy.deepcopy(pi.get("declaration_elements") or {}),
    }


def _product_block_from_excel(row: MappingRow) -> Dict[str, Any]:
    product_name = str(row.get("product_name") or "").strip()
    return {
        "hs_code": str(row.get("hs_code") or "").strip(),
        "product_name": product_name,
        "declaration_elements": copy.deepcopy(
            row.get("declaration_elements") or {"品名": product_name}
        ),
    }


def _match_internal_item_row(
    target_item: Any, internal_item_rows: List[MappingRow]
) -> Optional[MappingRow]:
    target = str(target_item or "").strip()
    if not target:
        return None
    target_l = target.lower()
    target_norm = _normalize_prefix_token(target)

    for row in internal_item_rows:
        item_name = str(row.get("item") or "").strip()
        if not item_name:
            continue
        if item_name.lower() == target_l:
            return row

        tokens = _split_rule_tokens(item_name)
        if any(token.lower() == target_l for token in tokens):
            return row

        norm_tokens = [_normalize_prefix_token(token) for token in tokens]
        if target_norm and target_norm in norm_tokens:
            return row
    return None


def _product_block_from_rule_target(
    loader: ConfigLoader,
    target_item: Any,
    internal_item_rows: List[MappingRow],
) -> Dict[str, Any]:
    target = str(target_item or "").strip()
    if not target:
        return {"hs_code": "", "product_name": "", "declaration_elements": {}}

    pi = loader.get_product_info(target)
    if pi:
        return _product_block_from_yaml(loader, target)

    row = _match_internal_item_row(target, internal_item_rows)
    if row:
        return _product_block_from_excel(row)

    return {"hs_code": "", "product_name": "", "declaration_elements": {}}


def _apply_leather_area_to_block(
    block: Dict[str, Any], item: Dict[str, Any]
) -> Dict[str, Any]:
    try:
        area = float(str(item.get("area_sqm") or "").replace(",", "."))
        count = float(str(item.get("hide_count") or item.get("quantity") or "").replace(",", "."))
    except (TypeError, ValueError):
        return block
    if area <= 0 or count <= 0:
        return block

    avg = area / count
    avg_text = _fmt_percent(avg)
    decl = block.get("declaration_elements") or {}
    for key, value in list(decl.items()):
        if "规格" not in str(key):
            continue
        text = "" if value is None else str(value)
        if "平均面积" in text:
            text = re.sub(r"平均面积[:：]\s*[^，,;；]*平方米", f"平均面积：{avg_text}平方米", text)
            if "平均面积：" not in text:
                text = f"{text}，平均面积：{avg_text}平方米"
        else:
            text = f"{text}，平均面积：{avg_text}平方米" if text else f"平均面积：{avg_text}平方米"
        decl[key] = text
    return block


def _with_cn_name(
    block: Dict[str, Any], name_cn: str
) -> Dict[str, Any]:
    out = {
        "hs_code": block["hs_code"],
        "product_name": name_cn,
        "declaration_elements": copy.deepcopy(block["declaration_elements"]),
    }
    decl = out["declaration_elements"]
    if name_cn:
        decl["品名"] = name_cn
    return out


def _item_desc_for_declaration(item: Dict[str, Any]) -> str:
    """Alcantara：品名列用 Material；申报/ITEM 匹配时拼接可选的 Description 栏。"""
    sup = (item.get("description_supplement") or "").strip()
    mat = (item.get("description") or "").strip()
    return f"{sup} {mat}".strip() if sup else mat


def _fmt_percent(value: float) -> str:
    if abs(value - round(value)) < 0.001:
        return str(int(round(value)))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def _parse_west_fabric_percentages(text: str) -> Dict[str, float]:
    values: Dict[str, float] = {}
    if not text:
        return values

    patterns = [
        r"(\d+(?:[,.]\d+)?)\s*%?\s*(wool|pes|polyester)\b",
        r"\b(wool|pes|polyester)\s*(\d+(?:[,.]\d+)?)\s*%",
    ]
    for pat in patterns:
        for m in re.finditer(pat, text, re.IGNORECASE):
            if m.group(1).lower() in {"wool", "pes", "polyester"}:
                material = m.group(1).lower()
                raw_num = m.group(2)
            else:
                raw_num = m.group(1)
                material = m.group(2).lower()
            try:
                pct = float(raw_num.replace(",", "."))
            except ValueError:
                continue
            key = "pes" if material in {"pes", "polyester"} else "wool"
            values.setdefault(key, pct)

    if values.get("wool") == 100 and "pes" not in values:
        values["pes"] = 0
    if values.get("pes") == 100 and "wool" not in values:
        values["wool"] = 0
    return values


def _composition_cn(percentages: Dict[str, float]) -> str:
    parts: List[str] = []
    wool = percentages.get("wool")
    pes = percentages.get("pes")
    if wool is not None and wool > 0:
        parts.append(f"{_fmt_percent(wool)}%羊毛")
    if pes is not None and pes > 0:
        parts.append(f"{_fmt_percent(pes)}%涤纶")
    return "，".join(parts)


def _is_west_trading(invoice_data: Optional[Dict[str, Any]]) -> bool:
    inv = invoice_data or {}
    supplier_code = str(inv.get("supplier_code") or "").strip().lower()
    supplier_name = str(inv.get("supplier_name") or "").strip().lower()
    return supplier_code == "west_trading" or "west trading" in supplier_name


def _west_trading_fabric_block(
    item: Dict[str, Any],
    loader: ConfigLoader,
    invoice_data: Optional[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    if not _is_west_trading(invoice_data):
        return None

    text = " ".join(
        str(v or "")
        for v in [
            item.get("item_code"),
            item.get("description"),
            item.get("description_supplement"),
            item.get("composition"),
            _item_desc_for_declaration(item),
        ]
    )
    percentages = _parse_west_fabric_percentages(text)
    wool = percentages.get("wool")
    pes = percentages.get("pes")
    if wool is None and pes is None:
        return None

    if pes is not None and pes >= 85:
        base_key = "west_pes_ge85"
    elif wool is not None and wool >= 85:
        base_key = "west_wool_ge85"
    elif wool is not None:
        base_key = "west_wool_lt85"
    else:
        return None

    block = _product_block_from_yaml(loader, base_key)
    block["product_name"] = "面料"

    composition = _composition_cn(percentages)
    if composition:
        decl = block["declaration_elements"]
        replaced = False
        for key in list(decl.keys()):
            if "成分" in str(key):
                decl[key] = composition
                replaced = True
        if not replaced:
            decl["成分"] = composition
    return block


def resolve_item_declaration(
    item: Dict[str, Any],
    excel_rows: List[MappingRow],
    loader: ConfigLoader,
    product_info: Optional[Dict[str, Any]],
    product_info_map: Dict[str, Dict[str, Any]],
    import_rules: Optional[List[MappingRow]] = None,
    invoice_data: Optional[Dict[str, Any]] = None,
    internal_item_rows: Optional[List[MappingRow]] = None,
) -> Dict[str, Any]:
    """
    返回单行的 hs_code, product_name, declaration_elements（已深拷贝）。
    """
    desc = _item_desc_for_declaration(item)
    pi_default = product_info or {}

    # 1) West Trading: 按每行 Wool/PES 实际比例动态归类并回填申报成分。
    west_fabric = _west_trading_fabric_block(item, loader, invoice_data)
    if west_fabric:
        return west_fabric

    # 2) 进口规则表：供应商原始描述/料号 -> YAML 内部标准品类
    if import_rules:
        rule = _match_import_rule(item, import_rules, invoice_data or {}, loader)
        if rule:
            blk = _product_block_from_rule_target(
                loader,
                rule.get("target_item") or rule.get("product_key"),
                internal_item_rows or [],
            )
            if blk.get("hs_code") or blk.get("declaration_elements"):
                blk = _apply_leather_area_to_block(blk, item)
                name_override = str(rule.get("product_name") or "").strip()
                return _with_cn_name(blk, name_override) if name_override else blk

    # 3) Excel ITEM 表（出口，或旧两列表兼容）
    ex = _match_excel_row(item, excel_rows)
    if ex:
        if ex.get("hs_code"):
            return _product_block_from_excel(ex)
        item_en = str(ex.get("item") or "").strip()
        name_cn = str(ex.get("product_name") or "").strip()
        yk = EXCEL_ITEM_TO_YAML_KEY.get(item_en)
        if yk:
            blk = _product_block_from_yaml(loader, yk)
            if blk.get("hs_code") or blk.get("declaration_elements"):
                return _with_cn_name(blk, name_cn)

    # 4) hide_type（与旧逻辑一致，在数字/MF/皮革规则之前便于专用皮料）
    hide_type = (item.get("hide_type") or "").strip()
    if hide_type and hide_type in product_info_map:
        cur = product_info_map[hide_type]
        block = {
            "hs_code": (cur.get("hs_code") or "").strip(),
            "product_name": (cur.get("name") or "").strip(),
            "declaration_elements": copy.deepcopy(
                cur.get("declaration_elements") or {}
            ),
        }
        return _apply_leather_area_to_block(block, item)

    # 5) item 前四位 5012 / 5015 → 面料 5603149000
    d4 = _first_four_digits(item.get("item_code"))
    if d4 == "5012":
        blk = _product_block_from_yaml(loader, "alcantara_5012")
        return _with_cn_name(blk, "面料")
    if d4 == "5015":
        blk = _product_block_from_yaml(loader, "alcantara_5015_5030_5010")
        return _with_cn_name(blk, "面料")

    # 6) MF 开头 → 超细纤维革 5903209000
    if _mf_prefix(desc):
        blk = _product_block_from_yaml(loader, "export_microfibre_hs5903209000")
        return _with_cn_name(blk, blk["product_name"] or "超细纤维革")

    # 7) NAPPA / VERONA / ROMA → 牛皮 4107121090
    if _leather_keyword_prefix(desc):
        blk = _product_block_from_yaml(loader, "mabo_cowhide")
        return _with_cn_name(blk, "牛皮")

    # 8) 默认 invoice 级 product_info
    return {
        "hs_code": (pi_default.get("hs_code") or "").strip(),
        "product_name": (pi_default.get("name") or "").strip(),
        "declaration_elements": copy.deepcopy(
            pi_default.get("declaration_elements") or {}
        ),
    }


def build_declaration_groups(
    items: List[Dict[str, Any]],
    invoice_data: Dict[str, Any],
    product_info: Optional[Dict[str, Any]],
    product_info_map: Optional[Dict[str, Dict[str, Any]]],
    config_path: str,
    mapping_xlsx: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """
    使用 item/品名 对应表与回落规则分组，结构同 DeclarationElementsGenerator._group_items_by_product。
    若未启用或缺少映射表，调用方应改用旧 _group_items_by_product。
    """
    cfg_path = _resolve_config_path(config_path)
    if not cfg_path.exists():
        cfg_path = ROOT / "data" / "supplier_product_mapping_import.yaml"
        if not cfg_path.exists():
            cfg_path = ROOT / "data" / "supplier_product_mapping.yaml"
    loader = ConfigLoader(str(cfg_path))

    is_export_flow = _is_export_flow(str(cfg_path), invoice_data)
    if is_export_flow:
        xlsx = _resolve_mapping_xlsx_path(str(cfg_path), mapping_xlsx, invoice_data)
        excel_rows = load_excel_mapping_rows(xlsx)
        import_rules: List[MappingRow] = []
    else:
        excel_rows = []
        import_rules = load_import_rules_for_config(
            str(cfg_path), invoice_data, mapping_xlsx
        )
    internal_items_xlsx = (
        _resolve_mapping_xlsx_path(str(cfg_path), None, invoice_data)
        if not is_export_flow
        else DEFAULT_EXPORT_EXCEL
    )
    internal_item_rows = load_excel_mapping_rows(internal_items_xlsx)

    pim = product_info_map or {}
    if not items:
        pi = product_info or {}
        return [
            {
                "hs_code": pi.get("hs_code", ""),
                "product_name": pi.get("name", ""),
                "declaration_elements": copy.deepcopy(
                    pi.get("declaration_elements") or {}
                ),
                "items": [],
            }
        ]

    groups: Dict[Tuple[Any, ...], Dict[str, Any]] = {}

    for item in items:
        resolved = resolve_item_declaration(
            item,
            excel_rows,
            loader,
            product_info,
            pim,
            import_rules=import_rules,
            invoice_data=invoice_data,
            internal_item_rows=internal_item_rows,
        )
        decl_key = tuple(
            (str(k), "" if v is None else str(v))
            for k, v in (resolved.get("declaration_elements") or {}).items()
        )
        key = (resolved["hs_code"], resolved["product_name"], decl_key)
        if key not in groups:
            groups[key] = {
                "hs_code": resolved["hs_code"],
                "product_name": resolved["product_name"],
                "declaration_elements": resolved["declaration_elements"],
                "items": [],
            }
        groups[key]["items"].append(item)

    return list(groups.values())


def _is_export_flow(
    config_path: str, invoice_data: Optional[Dict[str, Any]] = None
) -> bool:
    if (invoice_data or {}).get("mapping_flow") == "export":
        return True
    return "export" in _resolve_config_path(config_path).stem.lower()


def _resolve_mapping_xlsx_path(
    config_path: str,
    mapping_xlsx: Optional[str],
    invoice_data: Optional[Dict[str, Any]],
) -> Path:
    ovr = (invoice_data or {}).get("item_mapping_xlsx") or mapping_xlsx
    if ovr:
        p = Path(ovr)
        return p.resolve() if p.is_absolute() else (Path.cwd() / p).resolve()
    cfg = _resolve_config_path(config_path)
    if "export" in cfg.stem.lower() and DEFAULT_EXPORT_EXCEL.exists():
        return DEFAULT_EXPORT_EXCEL
    default_import = cfg.parent / "item和品名对应表.xlsx"
    return default_import if default_import.exists() else DEFAULT_IMPORT_EXCEL


def _resolve_import_rules_xlsx_path(
    config_path: str,
    mapping_xlsx: Optional[str],
    invoice_data: Optional[Dict[str, Any]],
) -> Path:
    ovr = (invoice_data or {}).get("import_rules_xlsx") or mapping_xlsx
    if ovr:
        p = Path(ovr)
        return p.resolve() if p.is_absolute() else (Path.cwd() / p).resolve()
    cfg = _resolve_config_path(config_path)
    default_import = cfg.parent / "进口商品识别规则.xlsx"
    return default_import if default_import.exists() else DEFAULT_IMPORT_RULES_EXCEL


def use_item_mapping_enabled(invoice_data: Dict[str, Any]) -> bool:
    """默认启用 ITEM/品名映射与回落规则；invoice_data['use_item_mapping']==False 时退回旧分组。"""
    return invoice_data.get("use_item_mapping", True) is not False
