"""
从 templates/CONTRACT.xlsx 生成 CONTRACT_CAMARI_TEMPLATE.xlsx
- 买方固定为 CAMARI TRADING (ZHEJIANG) CO., LTD
- 第12行：Currency / Incoterms / Payment terms（不插入行，避免破坏合并单元格）
- 保留 (5)-(8) 条款及签章区
"""
from pathlib import Path
import shutil
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "templates" / "CONTRACT.xlsx"
DST = ROOT / "templates" / "CONTRACT_CAMARI_TEMPLATE.xlsx"


def main():
    shutil.copy(SRC, DST)
    wb = load_workbook(DST)
    ws = wb.active

    ws.cell(row=4, column=3).value = "CAMARI TRADING (ZHEJIANG) CO., LTD"

    # 汇总行（表头上一行）：货币 / 贸易术语 / 付款条件（由程序填写）
    ws.cell(row=12, column=2).value = "Currency:"
    ws.cell(row=12, column=3).value = None
    ws.cell(row=12, column=5).value = "Incoterms:"
    ws.cell(row=12, column=6).value = None
    ws.cell(row=12, column=8).value = "Payment terms:"
    ws.cell(row=12, column=9).value = None
    # 付款条件可能较长，合并 I12:K12 作为填写区
    try:
        ws.merge_cells("I12:K12")
    except ValueError:
        pass

    # 清除误识别单元格
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == -11 or cell.value == "-11":
                cell.value = None

    # 模板空白：合同号、卖方、日期、首行单价列中的 EXW 占位
    ws.cell(row=5, column=11).value = None
    ws.cell(row=7, column=3).value = None
    ws.cell(row=7, column=11).value = None
    if ws.cell(row=15, column=10).value == "EXW":
        ws.cell(row=15, column=10).value = None

    wb.save(DST)
    print(f"Written: {DST}")


if __name__ == "__main__":
    main()
