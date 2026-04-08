#!/usr/bin/env python3
"""
Split a combined Excel file into separate Invoice and Packing List workbooks.

The input file is expected to contain sheets named:
  - "INV"  : invoice
  - "PACK" : packing list

This script preserves cell values, basic formatting (styles), merges, column widths,
and row heights to keep the layout "beautiful".
"""

from __future__ import annotations

import argparse
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries


def _copy_worksheet(src_ws, dest_wb: Workbook, title: str) -> None:
    dest_ws = dest_wb.active
    dest_ws.title = title

    dim = src_ws.calculate_dimension()  # e.g. "A1:I34"
    if not dim:
        return
    min_col, min_row, max_col, max_row = range_boundaries(dim)

    # Copy merged cells first (they need underlying cells existing).
    for merged in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(merged))

    # Copy cell values + styles.
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            src_cell = src_ws.cell(row=r, column=c)
            dest_cell = dest_ws.cell(row=r, column=c, value=src_cell.value)
            # Copy style only when it matters (value exists or style assigned).
            if src_cell.value is not None or src_cell.has_style:
                dest_cell._style = copy(src_cell._style)
                # Some style properties might not be in _style proxy copy reliably.
                dest_cell.number_format = src_cell.number_format
                dest_cell.alignment = copy(src_cell.alignment)
                dest_cell.font = copy(src_cell.font)
                dest_cell.border = copy(src_cell.border)
                dest_cell.fill = copy(src_cell.fill)
                dest_cell.protection = copy(src_cell.protection)

    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dest_ws.column_dimensions[col_letter].width = dim.width

    # Copy row heights
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dest_ws.row_dimensions[row_idx].height = dim.height

    # Copy view settings (best-effort)
    try:
        dest_ws.sheet_view = copy(src_ws.sheet_view)
    except Exception:
        pass

    # Copy print settings (best-effort)
    try:
        dest_ws.page_setup = copy(src_ws.page_setup)
    except Exception:
        pass


def split(input_xlsx: str, output_dir: str) -> tuple[Path, Path]:
    in_path = Path(input_xlsx).expanduser().resolve()
    out_dir = Path(output_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(in_path, data_only=True)
    sheet_inv = "INV"
    sheet_pack = "PACK"

    if sheet_inv not in wb.sheetnames or sheet_pack not in wb.sheetnames:
        raise ValueError(
            f"Expected sheets {sheet_inv} and {sheet_pack}. Found: {wb.sheetnames}"
        )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = in_path.stem

    inv_out = out_dir / f"invoice_{stem}_{ts}.xlsx"
    pack_out = out_dir / f"packing_list_{stem}_{ts}.xlsx"

    dest_inv = Workbook()
    # Remove the default sheet created by Workbook()
    default = dest_inv.active
    default.title = "INV"
    _copy_worksheet(wb[sheet_inv], dest_inv, "INV")
    # If copy changed active sheet unexpectedly, ensure only INV exists
    # (we don't over-optimize; layout is the priority)
    dest_inv.save(inv_out)

    dest_pack = Workbook()
    default = dest_pack.active
    default.title = "PACK"
    _copy_worksheet(wb[sheet_pack], dest_pack, "PACK")
    dest_pack.save(pack_out)

    return inv_out, pack_out


def main():
    ap = argparse.ArgumentParser(description="Split INV/PACK sheets into separate Excel files.")
    ap.add_argument("xlsx", help="Input xlsx path (e.g. 发票箱单.xlsx)")
    ap.add_argument(
        "-o",
        "--output",
        default="output",
        help="Output directory (default: output)",
    )
    args = ap.parse_args()

    inv_out, pack_out = split(args.xlsx, args.output)
    print(f"Invoice saved: {inv_out}")
    print(f"Packing list saved: {pack_out}")


if __name__ == "__main__":
    main()

